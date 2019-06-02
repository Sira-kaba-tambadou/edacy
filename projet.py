import os
import openpyxl
from openpyxl import *
from openpyxl.utils import get_column_letter
 
wb = Workbook()
wb= openpyxl.load_workbook('test.xlsx')
 
# On active l'onglet courant
ws = wb.active
 
# On crée un nouvel onglet
ws1 = wb.create_sheet()
ws1.title = ws.title
 
# On s'apprête à parcourir la colonne B
translated = []
begrow = 1
endrow = ws.max_row
for row in range(begrow,endrow):
    endrows1 = ws1.max_row + 1
    for col in range(1, 7):
        print (endrows1)
        if ws['B' + format(row)].value is not None:
            ws1.cell(column=col, row=endrows1, value=ws[format(get_column_letter(col)) + format(row)].value)
 
wb.save('file.xlsx')
wb.close()
